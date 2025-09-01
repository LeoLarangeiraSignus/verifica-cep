import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import json


def verifica_cep(ceps):

    cep = ceps["cep"]
    email = ceps["email"].strip()
    nome = ceps["nome"].strip()
    url = f"https://viacep.com.br/ws/{cep}/json/"

    try:
        response = requests.get(url, timeout=20)
        if response.status_code == 200:
            data = response.json()
            if "erro" in data:  # CEP inválido
                cria_excel(
                    "SA1", cep, email, nome, "Formato Válido, porém CEP inexistente"
                )
        else:
            print(response)
            cria_excel("SA1", cep, email, nome, "Formato inválido")

    except requests.RequestException as e:
        print(f"Erro na requisição HTTP: {e}")


def cria_excel(table_name, cep, email, nome, motivo):
    filename = "ceps-errados.xlsx"
    if os.path.isfile(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        last_row = sheet.max_row + 1
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"CEPs Errados tabela {table_name}"

        # Cabeçalhos
        sheet["A1"] = "Email"
        sheet["B1"] = "CEP"
        sheet["C1"] = "Nome"
        sheet["D1"] = "Motivo"

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")
        header_fill = PatternFill(
            start_color="C8C8C8", end_color="C8C8C8", fill_type="solid"
        )

        for col in ["A1", "B1", "C1", "D1"]:
            sheet[col].font = header_font
            sheet[col].alignment = header_alignment
            sheet[col].fill = header_fill

        last_row = 2  # Primeira linha de dados

    # Escreve os dados
    sheet[f"A{last_row}"] = email
    sheet[f"B{last_row}"] = cep
    sheet[f"C{last_row}"] = nome
    sheet[f"D{last_row}"] = motivo

    # Ajusta largura das colunas
    for col in ["A", "B", "C", "D"]:
        max_length = 0
        column = sheet[col]
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[col].width = adjusted_width

    workbook.save(filename)
    workbook.close()
